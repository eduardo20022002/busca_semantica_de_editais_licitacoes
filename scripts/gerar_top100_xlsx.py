"""Gera um .xlsx com o Top 100 (Score V4) enriquecido com dados completos do edital.

Uso (rodar da raiz do repo):
    uv run python scripts/gerar_top100_xlsx.py
"""
import glob
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parent.parent

top = json.loads((RAIZ / "base-teste/comparacao-v1-v2-v3/top100_v4.json").read_text(encoding="utf-8"))

indice = {}
for caminho in sorted(glob.glob(str(RAIZ / "dados/editais_2026-07-*.json"))):
    for e in json.loads(Path(caminho).read_text(encoding="utf-8")):
        indice[e["numeroControlePNCP"]] = e

FONTE = "Arial"

wb = Workbook()
ws = wb.active
ws.title = "Top 100 - Score V4"

colunas = [
    ("Posição", 8),
    ("Score V4", 11),
    ("Produto Aurora mais próximo", 30),
    ("Objeto da compra", 60),
    ("Órgão / Entidade", 34),
    ("Município", 20),
    ("UF", 6),
    ("Modalidade", 22),
    ("Valor estimado (R$)", 18),
    ("Data abertura proposta", 16),
    ("Data encerramento proposta", 16),
    ("Situação", 20),
    ("Número controle PNCP", 26),
    ("Link sistema origem", 40),
]

fonte_cabecalho = Font(name=FONTE, bold=True, color="FFFFFF", size=10)
preenchimento_cabecalho = PatternFill("solid", fgColor="1F4E78")
fonte_corpo = Font(name=FONTE, size=10)
alinhamento_topo = Alignment(vertical="top", wrap_text=True)
borda_fina = Border(bottom=Side(style="thin", color="D9D9D9"))

for col_idx, (titulo, largura) in enumerate(colunas, start=1):
    celula = ws.cell(row=1, column=col_idx, value=titulo)
    celula.font = fonte_cabecalho
    celula.fill = preenchimento_cabecalho
    celula.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(col_idx)].width = largura

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30

for i, r in enumerate(top, start=1):
    linha = i + 1
    numero = r["numeroControlePNCP"]
    completo = indice.get(numero, {})

    orgao = completo.get("orgaoEntidade") or {}
    unidade = completo.get("unidadeOrgao") or {}

    valores = [
        i,
        round(r["score_v4"], 4),
        r["produto_v4"],
        r["objetoCompra"].strip(),
        orgao.get("razaoSocial", ""),
        unidade.get("municipioNome", ""),
        unidade.get("ufSigla", ""),
        completo.get("modalidadeNome", ""),
        completo.get("valorTotalEstimado"),
        (completo.get("dataAberturaProposta") or "")[:10],
        (completo.get("dataEncerramentoProposta") or "")[:10],
        completo.get("situacaoCompraNome", ""),
        numero,
        completo.get("linkSistemaOrigem", ""),
    ]

    for col_idx, valor in enumerate(valores, start=1):
        celula = ws.cell(row=linha, column=col_idx, value=valor)
        celula.font = fonte_corpo
        celula.alignment = alinhamento_topo
        celula.border = borda_fina

    ws.cell(row=linha, column=2).number_format = "0.0000"
    ws.cell(row=linha, column=9).number_format = "R$ #,##0.00;(R$ #,##0.00);-"

    link_url = completo.get("linkSistemaOrigem")
    if link_url:
        celula_link = ws.cell(row=linha, column=14)
        celula_link.hyperlink = link_url
        celula_link.font = Font(name=FONTE, size=10, color="0563C1", underline="single")

ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{len(top) + 1}"

# --- segunda aba: resumo por produto (via fórmulas, não hardcoded) ---
resumo = wb.create_sheet("Resumo por Produto")
cabecalhos_resumo = ["Produto Aurora", "Qtd no Top 100", "Score médio", "Score máximo"]
for col_idx, titulo in enumerate(cabecalhos_resumo, start=1):
    celula = resumo.cell(row=1, column=col_idx, value=titulo)
    celula.font = fonte_cabecalho
    celula.fill = preenchimento_cabecalho
    celula.alignment = Alignment(horizontal="center")
resumo.column_dimensions["A"].width = 34
resumo.column_dimensions["B"].width = 16
resumo.column_dimensions["C"].width = 14
resumo.column_dimensions["D"].width = 14

produtos = sorted({r["produto_v4"] for r in top})
ref = "'Top 100 - Score V4'"
for idx, produto in enumerate(produtos, start=2):
    resumo.cell(row=idx, column=1, value=produto).font = fonte_corpo
    resumo.cell(row=idx, column=2, value=f"=COUNTIF({ref}!C:C,A{idx})").font = fonte_corpo
    resumo.cell(row=idx, column=3, value=f'=IFERROR(AVERAGEIF({ref}!C:C,A{idx},{ref}!B:B),"-")').font = fonte_corpo
    resumo.cell(row=idx, column=4, value=f'=IFERROR(_xlfn.MAXIFS({ref}!B:B,{ref}!C:C,A{idx}),"-")').font = fonte_corpo
    resumo.cell(row=idx, column=3).number_format = "0.0000"
    resumo.cell(row=idx, column=4).number_format = "0.0000"

linha_total = len(produtos) + 2
resumo.cell(row=linha_total, column=1, value="Total").font = Font(name=FONTE, bold=True, size=10)
resumo.cell(row=linha_total, column=2, value=f"=SUM(B2:B{linha_total - 1})").font = Font(name=FONTE, bold=True, size=10)

destino = RAIZ / "base-teste/comparacao-v1-v2-v3/top100_v4.xlsx"
wb.save(destino)
print("Salvo em", destino)
